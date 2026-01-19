"""
Audit Script: Compare flood investigation data between source and destination

This script compares folders between:
    Source: I:\\FloodInvestigations
    Destination: data\\flood (relative to project)

Generates:
    audit.xlsx - Summary of layer types per area (height, depths, velocity)
    audit_details.xlsx - Detailed year-by-year breakdown for each layer type

Usage:
    python audit.py
"""

import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

# Paths
SOURCE_DIR = Path(r"I:\FloodInvestigations")
DEST_DIR = Path(r"C:\Users\m.rahman\qgis\pozi_mirror\data\flood")
SCRIPT_DIR = Path(__file__).parent

# Layer types to check
LAYER_TYPES = ["Height", "Depths", "Velocity"]

# Years to check
YEARS = [5, 10, 20, 50, 100, 200]

# Fonts
RED_FONT = Font(color="FF0000")
BOLD_FONT = Font(bold=True)


def get_all_study_folders():
    """Get all study folder names from destination."""
    if not DEST_DIR.exists():
        return []
    
    folders = []
    for item in DEST_DIR.iterdir():
        if item.is_dir():
            folders.append(item.name)
    
    return sorted(folders)


def folder_exists(area: str, layer_type: str) -> bool:
    """Check if a layer type folder exists for an area."""
    folder_path = DEST_DIR / area / layer_type
    return folder_path.exists() and folder_path.is_dir()


def year_file_exists(area: str, layer_type: str, year: int) -> bool:
    """
    Check if a file for a specific year exists in the layer type folder.
    
    Matches year precisely (e.g., 10 should not match 100).
    """
    folder_path = DEST_DIR / area / layer_type
    
    if not folder_path.exists():
        return False
    
    # Pattern to match year precisely (not part of larger number)
    # Examples: _10y_, _10_, ARI10, Vel10, etc.
    year_pattern = re.compile(rf'(?<!\d){year}(?!\d)')
    
    for file in folder_path.iterdir():
        if file.is_file() and file.suffix.lower() in ['.tif', '.shp']:
            if year_pattern.search(file.stem):
                return True
    
    return False


def generate_audit_xlsx(areas: list):
    """Generate audit.xlsx with summary of layer types per area."""
    output_file = SCRIPT_DIR / "audit.xlsx"
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Audit"
    
    # Header
    headers = ["area", "height", "depths", "velocity"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = BOLD_FONT
    
    # Data rows
    for row_idx, area in enumerate(areas, 2):
        ws.cell(row=row_idx, column=1, value=area)
        
        for col_idx, layer_type in enumerate(LAYER_TYPES, 2):
            exists = folder_exists(area, layer_type)
            cell = ws.cell(row=row_idx, column=col_idx, value="Yes" if exists else "No")
            if not exists:
                cell.font = RED_FONT
    
    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_length + 2
    
    wb.save(output_file)
    print(f"Generated: {output_file}")
    return output_file


def generate_audit_details_xlsx(areas: list):
    """Generate audit_details.xlsx with year-by-year breakdown."""
    output_file = SCRIPT_DIR / "audit_details.xlsx"
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Audit Details"
    
    # Build header: area, height_5, height_10, ..., velocity_200
    headers = ["area"]
    for layer_type in LAYER_TYPES:
        for year in YEARS:
            headers.append(f"{layer_type.lower()}_{year}")
    
    # Write header
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = BOLD_FONT
    
    # Data rows
    for row_idx, area in enumerate(areas, 2):
        ws.cell(row=row_idx, column=1, value=area)
        
        col_idx = 2
        for layer_type in LAYER_TYPES:
            for year in YEARS:
                exists = year_file_exists(area, layer_type, year)
                cell = ws.cell(row=row_idx, column=col_idx, value="Yes" if exists else "No")
                if not exists:
                    cell.font = RED_FONT
                col_idx += 1
    
    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_length + 2
    
    wb.save(output_file)
    print(f"Generated: {output_file}")
    return output_file


def audit():
    """Main audit function."""
    print("=" * 70)
    print("AUDIT: Flood Data Audit")
    print("=" * 70)
    print(f"\nDestination: {DEST_DIR}")
    
    # Check directory exists
    if not DEST_DIR.exists():
        print(f"\nERROR: Destination directory not found: {DEST_DIR}")
        return
    
    # Get all study folders
    areas = get_all_study_folders()
    print(f"Found {len(areas)} area(s)")
    
    # Generate Excel files
    print("\nGenerating Excel reports...")
    generate_audit_xlsx(areas)
    generate_audit_details_xlsx(areas)
    
    # Print summary statistics
    print("\n" + "=" * 70)
    print("SUMMARY")
    print("=" * 70)
    
    # Count completeness
    complete_folders = 0
    complete_details = 0
    
    for area in areas:
        # Check folder completeness
        all_folders = all(folder_exists(area, lt) for lt in LAYER_TYPES)
        if all_folders:
            complete_folders += 1
        
        # Check year completeness
        all_years = all(
            year_file_exists(area, lt, y)
            for lt in LAYER_TYPES
            for y in YEARS
        )
        if all_years:
            complete_details += 1
    
    print(f"\n  Total areas: {len(areas)}")
    print(f"  Areas with all 3 layer type folders: {complete_folders}")
    print(f"  Areas with all years in all layer types: {complete_details}")
    
    print("\n" + "=" * 70)
    print("Audit completed!")
    print("=" * 70)


if __name__ == "__main__":
    audit()
